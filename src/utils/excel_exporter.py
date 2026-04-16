# src/utils/excel_exporter.py
"""
Модуль для экспорта данных в Excel (XLSX)
Использует openpyxl для создания форматированных отчётов
"""

import os
from datetime import datetime, date
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """
    Экспорт данных в Excel с форматированием
    """
    
    # Цвета
    HEADER_BG = "34495e"      # Тёмно-синий
    HEADER_FG = "FFFFFF"      # Белый текст
    SUCCESS_BG = "d5f4e6"     # Светло-зелёный (для выполненных)
    WARNING_BG = "fef9e7"     # Светло-жёлтый (в работе)
    INFO_BG = "e3f2fd"        # Светло-синий (очередь)
    DANGER_BG = "fdedec"      # Светло-красный (отменено)
    TOTAL_BG = "ecf0f1"       # Серый для итогов
    
    def __init__(self):
        self.wb = None
        self.header_font = Font(name='Arial', size=11, bold=True, color=self.HEADER_FG)
        self.header_fill = PatternFill(start_color=self.HEADER_BG, end_color=self.HEADER_BG, fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        self.thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
    
    def _apply_header_style(self, ws, row: int, columns: List[str]):
        """Применяет стиль заголовка к строке"""
        for col_idx, header in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
    
    def _apply_cell_style(self, cell, fill_color: str = None, bold: bool = False, 
                          alignment: str = 'left', number_format: str = None):
        """Применяет стиль к ячейке"""
        cell.border = self.thin_border
        cell.alignment = Alignment(horizontal=alignment, vertical='center')
        
        if bold:
            cell.font = Font(name='Arial', size=10, bold=True)
        else:
            cell.font = Font(name='Arial', size=10)
        
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        if number_format:
            cell.number_format = number_format
    
    def _auto_size_columns(self, ws):
        """Автоматически подбирает ширину колонок"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # ============ ЭКСПОРТ ЗАКАЗОВ ============
    
    def export_orders(
        self,
        orders: List[Dict[str, Any]],
        statistics: Optional[Dict[str, Any]] = None,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        filepath: str = None
    ) -> str:
        """
        Экспортирует заказы в Excel
        
        Args:
            orders: Список заказов
            statistics: Статистика по заказам
            date_from: Дата начала периода
            date_to: Дата конца периода
            filepath: Путь для сохранения (если None — генерируется)
            
        Returns:
            Путь к сохранённому файлу
        """
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "Заказы"
        
        # Заголовок отчёта
        row = 1
        ws.merge_cells(f'A{row}:I{row}')
        title_cell = ws.cell(row=row, column=1, value="📋 ОТЧЁТ ПО ЗАКАЗАМ")
        title_cell.font = Font(name='Arial', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Период
        if date_from and date_to:
            ws.merge_cells(f'A{row}:I{row}')
            period_cell = ws.cell(row=row, column=1, value=f"Период: {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}")
            period_cell.font = Font(name='Arial', size=11)
            period_cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
        
        # Дата формирования
        ws.merge_cells(f'A{row}:I{row}')
        date_cell = ws.cell(row=row, column=1, value=f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        date_cell.font = Font(name='Arial', size=10, italic=True)
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Статистика (если есть)
        if statistics:
            stats_start_row = row
            ws.cell(row=row, column=1, value="📊 СТАТИСТИКА").font = Font(name='Arial', size=12, bold=True)
            row += 1
            
            stats_data = [
                ("Всего заказов:", statistics.get('total_orders', 0)),
                ("Выручка:", f"{statistics.get('total_revenue', 0):,.0f} ₽".replace(',', ' ')),
            ]
            
            for label, value in stats_data:
                label_cell = ws.cell(row=row, column=1, value=label)
                label_cell.font = Font(name='Arial', size=10, bold=True)
                value_cell = ws.cell(row=row, column=2, value=value)
                value_cell.font = Font(name='Arial', size=10)
                row += 1
            
            row += 1
            
            # Оплата
            ws.cell(row=row, column=1, value="По способам оплаты:").font = Font(name='Arial', size=11, bold=True)
            row += 1
            
            payment_data = [
                ("💵 Наличные:", statistics.get('cash_amount', 0), statistics.get('cash_count', 0)),
                ("💳 Карта:", statistics.get('card_amount', 0), statistics.get('card_count', 0)),
                ("📱 Перевод:", statistics.get('transfer_amount', 0), statistics.get('transfer_count', 0)),
                ("📲 СБП:", statistics.get('sbp_amount', 0), statistics.get('sbp_count', 0)),
            ]
            
            for label, amount, count in payment_data:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=f"{amount:,.0f} ₽".replace(',', ' '))
                ws.cell(row=row, column=3, value=f"({count} зак.)")
                row += 1
            
            row += 1
        
        # Заголовки таблицы
        headers = ["ID", "Дата/Время", "Гос.номер", "Марка", "Телефон", "Услуги", "Статус", "Сумма", "Оплата"]
        self._apply_header_style(ws, row, headers)
        row += 1
        
        # Данные заказов
        status_colors = {
            'queue': self.INFO_BG,
            'process': self.WARNING_BG,
            'done': self.SUCCESS_BG,
            'cancelled': self.DANGER_BG
        }
        
        status_names = {
            'queue': 'В очереди',
            'process': 'В работе',
            'done': 'Готово',
            'cancelled': 'Отменено'
        }
        
        payment_names = {
            'cash': 'Наличные',
            'card': 'Карта',
            'transfer': 'Перевод',
            'sbp': 'СБП'
        }
        
        total_sum = 0
        
        for order in orders:
            status = order.get('status', 'queue')
            fill_color = status_colors.get(status, '')
            
            # ID
            cell = ws.cell(row=row, column=1, value=order.get('id'))
            self._apply_cell_style(cell, fill_color, alignment='center')
            
            # Дата/время
            date_str = order.get('created_at', '')
            if date_str:
                if isinstance(date_str, str):
                    date_str = date_str[:16]
            cell = ws.cell(row=row, column=2, value=date_str)
            self._apply_cell_style(cell, fill_color)
            
            # Гос.номер
            cell = ws.cell(row=row, column=3, value=order.get('car_number', '—'))
            self._apply_cell_style(cell, fill_color, bold=True)
            
            # Марка
            cell = ws.cell(row=row, column=4, value=order.get('car_model', '—'))
            self._apply_cell_style(cell, fill_color)
            
            # Телефон
            cell = ws.cell(row=row, column=5, value=order.get('client_phone', '—'))
            self._apply_cell_style(cell, fill_color)
            
            # Услуги
            cell = ws.cell(row=row, column=6, value=order.get('services_list', '—'))
            self._apply_cell_style(cell, fill_color)
            
            # Статус
            cell = ws.cell(row=row, column=7, value=status_names.get(status, status))
            self._apply_cell_style(cell, fill_color, alignment='center')
            
            # Сумма
            amount = order.get('total_price', 0) or 0
            total_sum += amount
            cell = ws.cell(row=row, column=8, value=amount)
            cell.number_format = '#,##0.00 ₽'
            self._apply_cell_style(cell, fill_color, bold=True, alignment='right')
            
            # Оплата
            payment = order.get('payment_method', '')
            cell = ws.cell(row=row, column=9, value=payment_names.get(payment, payment))
            self._apply_cell_style(cell, fill_color, alignment='center')
            
            row += 1
        
        # Итоговая строка
        row += 1
        ws.merge_cells(f'A{row}:G{row}')
        total_label = ws.cell(row=row, column=1, value="ИТОГО:")
        total_label.font = Font(name='Arial', size=11, bold=True)
        total_label.alignment = Alignment(horizontal='right', vertical='center')
        self._apply_cell_style(total_label, self.TOTAL_BG)
        
        total_cell = ws.cell(row=row, column=8, value=total_sum)
        total_cell.number_format = '#,##0.00 ₽'
        total_cell.font = Font(name='Arial', size=12, bold=True)
        total_cell.alignment = Alignment(horizontal='right', vertical='center')
        self._apply_cell_style(total_cell, self.TOTAL_BG)
        
        self._apply_cell_style(ws.cell(row=row, column=9), self.TOTAL_BG)
        
        # Автоширина колонок
        self._auto_size_columns(ws)
        
        # Сохраняем
        if filepath is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filepath = f"orders_export_{timestamp}.xlsx"
        
        self.wb.save(filepath)
        return filepath
    
    # ============ ЭКСПОРТ КЛИЕНТОВ ============
    
    def export_clients(
        self,
        clients: List[Dict[str, Any]],
        filepath: str = None
    ) -> str:
        """
        Экспортирует клиентов в Excel
        
        Args:
            clients: Список клиентов
            filepath: Путь для сохранения
            
        Returns:
            Путь к сохранённому файлу
        """
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "Клиенты"
        
        # Заголовок
        row = 1
        ws.merge_cells(f'A{row}:H{row}')
        title_cell = ws.cell(row=row, column=1, value="👤 БАЗА КЛИЕНТОВ")
        title_cell.font = Font(name='Arial', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Дата
        ws.merge_cells(f'A{row}:H{row}')
        date_cell = ws.cell(row=row, column=1, value=f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        date_cell.font = Font(name='Arial', size=10, italic=True)
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Статистика
        ws.cell(row=row, column=1, value=f"Всего клиентов: {len(clients)}").font = Font(name='Arial', size=11, bold=True)
        row += 2
        
        # Заголовки таблицы
        headers = ["ID", "Гос.номер", "Марка", "Телефон", "Визитов", "Потрачено", "Последний визит", "Комментарий"]
        self._apply_header_style(ws, row, headers)
        row += 1
        
        # Данные клиентов
        for client in clients:
            # ID
            cell = ws.cell(row=row, column=1, value=client.get('id'))
            self._apply_cell_style(cell, alignment='center')
            
            # Гос.номер
            cell = ws.cell(row=row, column=2, value=client.get('car_number', '—'))
            self._apply_cell_style(cell, bold=True)
            
            # Марка
            cell = ws.cell(row=row, column=3, value=client.get('car_model', '—'))
            self._apply_cell_style(cell)
            
            # Телефон
            cell = ws.cell(row=row, column=4, value=client.get('phone', '—'))
            self._apply_cell_style(cell)
            
            # Визитов
            visits = client.get('total_visits', 0) or 0
            cell = ws.cell(row=row, column=5, value=visits)
            self._apply_cell_style(cell, alignment='center')
            
            # Потрачено
            spent = client.get('total_spent', 0) or 0
            cell = ws.cell(row=row, column=6, value=spent)
            cell.number_format = '#,##0.00 ₽'
            self._apply_cell_style(cell, alignment='right')
            
            # Последний визит
            last_visit = client.get('last_visit', '')
            if last_visit and isinstance(last_visit, str):
                last_visit = last_visit[:10]
            cell = ws.cell(row=row, column=7, value=last_visit or '—')
            self._apply_cell_style(cell, alignment='center')
            
            # Комментарий
            cell = ws.cell(row=row, column=8, value=client.get('comment', '—'))
            self._apply_cell_style(cell)
            
            row += 1
        
        # Автоширина
        self._auto_size_columns(ws)
        
        # Сохраняем
        if filepath is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filepath = f"clients_export_{timestamp}.xlsx"
        
        self.wb.save(filepath)
        return filepath